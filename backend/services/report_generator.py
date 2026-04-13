"""
factory-ai-platform · services/report_generator.py
Generates machine & plant-wide reports in PDF and Excel formats.
"""
import io
import json
import logging
from datetime import datetime, timezone, timedelta
from typing import Any

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

from database.db import execute_many, execute_one, execute_write
from config.settings import settings

logger = logging.getLogger(__name__)

YASH_BLUE = colors.HexColor('#0057A8')
YASH_RED  = colors.HexColor('#E31837')
DARK      = colors.HexColor('#0F172A')
GRAY      = colors.HexColor('#64748B')
LIGHT     = colors.HexColor('#F8FAFC')


def _safe_float(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


class ReportGenerator:

    # ── PDF Report ────────────────────────────────────────────
    def generate_pdf(
        self,
        machine_id: int,
        report_type: str,
        start_dt: datetime,
        end_dt: datetime,
        generated_by: int | None = None,
    ) -> bytes:
        """Generate a PDF report and return as bytes."""
        machine = execute_one(
            "SELECT id, machine_name, location, machine_type FROM machines WHERE id = %s",
            (machine_id,),
        )
        if not machine:
            raise ValueError(f"Machine {machine_id} not found")

        stats = self._get_stats(machine_id, start_dt, end_dt)
        alerts = self._get_alerts(machine_id, start_dt, end_dt)
        telemetry = self._get_telemetry(machine_id, start_dt, end_dt)

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=A4,
            rightMargin=20*mm, leftMargin=20*mm,
            topMargin=20*mm, bottomMargin=20*mm,
        )

        styles = getSampleStyleSheet()
        story  = []

        # ── Header ──────────────────────────────────────────
        header_data = [[
            Paragraph(f'<font color="#0057A8"><b>EDGEAI</b></font><br/><font size="8" color="#64748B">Yash Technologies · Machine Report</font>', styles['Normal']),
            Paragraph(f'<font size="8" color="#64748B">Generated: {datetime.now().strftime("%d %b %Y, %H:%M")}</font>', ParagraphStyle('right', alignment=TA_LEFT)),
        ]]
        header_table = Table(header_data, colWidths=[120*mm, 50*mm])
        header_table.setStyle(TableStyle([
            ('BACKGROUND',  (0,0), (-1,-1), YASH_BLUE),
            ('TEXTCOLOR',   (0,0), (-1,-1), colors.white),
            ('PADDING',     (0,0), (-1,-1), 12),
            ('ROWBACKGROUNDS', (0,0), (-1,-1), [YASH_BLUE]),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 6*mm))

        # ── Title ────────────────────────────────────────────
        story.append(Paragraph(
            f'<font size="18" color="#0F172A"><b>{machine["machine_name"]} — {report_type.capitalize()} Report</b></font>',
            styles['Normal']
        ))
        story.append(Paragraph(
            f'<font size="10" color="#64748B">{machine["machine_type"]} · {machine["location"]} · {start_dt.strftime("%d %b %Y")} to {end_dt.strftime("%d %b %Y")}</font>',
            styles['Normal']
        ))
        story.append(Spacer(1, 5*mm))
        story.append(HRFlowable(width="100%", thickness=2, color=YASH_RED))
        story.append(Spacer(1, 5*mm))

        # ── Summary Stats ────────────────────────────────────
        story.append(Paragraph('<b>Performance Summary</b>', styles['Heading2']))
        story.append(Spacer(1, 3*mm))

        stats_data = [
            ['Parameter', 'Average', 'Maximum', 'Minimum', 'Status'],
            ['Temperature (°C)',
             f"{_safe_float(stats.get('avg_temp')):.1f}",
             f"{_safe_float(stats.get('max_temp')):.1f}",
             f"{_safe_float(stats.get('min_temp')):.1f}",
             '⚠ High' if _safe_float(stats.get('max_temp')) > 95 else '✓ Normal'],
            ['Vibration (mm/s)',
             f"{_safe_float(stats.get('avg_vibration')):.4f}",
             f"{_safe_float(stats.get('max_vibration')):.4f}",
             '—',
             '⚠ High' if _safe_float(stats.get('max_vibration')) > 0.4 else '✓ Normal'],
            ['RPM',
             f"{_safe_float(stats.get('avg_rpm')):.0f}",
             '—', '—',
             '✓ Normal'],
            ['Power (W)',
             f"{_safe_float(stats.get('avg_power')):.1f}",
             f"{_safe_float(stats.get('max_power')):.1f}",
             '—',
             '⚠ High' if _safe_float(stats.get('max_power')) > 1500 else '✓ Normal'],
        ]

        stats_table = Table(stats_data, colWidths=[45*mm, 30*mm, 30*mm, 30*mm, 30*mm])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND',   (0,0), (-1,0), YASH_BLUE),
            ('TEXTCOLOR',    (0,0), (-1,0), colors.white),
            ('FONTNAME',     (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',     (0,0), (-1,-1), 9),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [LIGHT, colors.white]),
            ('GRID',         (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('PADDING',      (0,0), (-1,-1), 7),
            ('ALIGN',        (1,0), (-1,-1), 'CENTER'),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 5*mm))

        # ── Key Metrics ──────────────────────────────────────
        total_readings  = int(stats.get('reading_count') or 0)
        anomaly_count   = int(stats.get('anomaly_count') or 0)
        uptime          = self._calc_uptime(machine_id, start_dt, end_dt)
        anomaly_rate    = round((anomaly_count / max(total_readings, 1)) * 100, 2)

        metrics_data = [
            ['Total Readings', 'Anomalies Detected', 'Anomaly Rate', 'Est. Uptime'],
            [str(total_readings), str(anomaly_count), f'{anomaly_rate}%', f'{uptime}%'],
        ]
        metrics_table = Table(metrics_data, colWidths=[42*mm, 42*mm, 42*mm, 42*mm])
        metrics_table.setStyle(TableStyle([
            ('BACKGROUND',  (0,0), (-1,0), colors.HexColor('#F1F5F9')),
            ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',    (0,0), (-1,-1), 9),
            ('ALIGN',       (0,0), (-1,-1), 'CENTER'),
            ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
            ('PADDING',     (0,0), (-1,-1), 8),
            ('FONTNAME',    (0,1), (-1,1), 'Helvetica-Bold'),
            ('FONTSIZE',    (0,1), (-1,1), 14),
            ('TEXTCOLOR',   (0,1), (-1,1), YASH_BLUE),
        ]))
        story.append(metrics_table)
        story.append(Spacer(1, 5*mm))

        # ── Alerts ───────────────────────────────────────────
        if alerts:
            story.append(Paragraph('<b>Alerts During Period</b>', styles['Heading2']))
            story.append(Spacer(1, 3*mm))
            alert_data = [['Timestamp', 'Type', 'Severity', 'Message']]
            for a in alerts[:15]:
                alert_data.append([
                    str(a['timestamp'])[:16] if a['timestamp'] else '—',
                    str(a['alert_type']).replace('_', ' ').title(),
                    str(a['severity']).upper(),
                    str(a['message'])[:60] + ('...' if len(str(a['message'])) > 60 else ''),
                ])
            alert_table = Table(alert_data, colWidths=[35*mm, 35*mm, 22*mm, 75*mm])
            alert_table.setStyle(TableStyle([
                ('BACKGROUND',  (0,0), (-1,0), YASH_RED),
                ('TEXTCOLOR',   (0,0), (-1,0), colors.white),
                ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTSIZE',    (0,0), (-1,-1), 8),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor('#FEF2F2'), colors.white]),
                ('GRID',        (0,0), (-1,-1), 0.5, colors.HexColor('#E2E8F0')),
                ('PADDING',     (0,0), (-1,-1), 6),
            ]))
            story.append(alert_table)
            story.append(Spacer(1, 5*mm))

        # ── AI Summary ───────────────────────────────────────
        summary = self._rule_based_summary(stats, alerts, machine)
        story.append(Paragraph('<b>AI Summary & Recommendations</b>', styles['Heading2']))
        story.append(Spacer(1, 3*mm))
        story.append(Paragraph(summary.replace('\n', '<br/>'), styles['Normal']))
        story.append(Spacer(1, 5*mm))

        # ── Footer ───────────────────────────────────────────
        story.append(HRFlowable(width="100%", thickness=1, color=GRAY))
        story.append(Spacer(1, 2*mm))
        story.append(Paragraph(
            f'<font size="8" color="#94A3B8">EDGEAI · Yash Technologies · Report generated on {datetime.now().strftime("%d %b %Y at %H:%M")} · Confidential</font>',
            ParagraphStyle('footer', alignment=TA_CENTER)
        ))

        doc.build(story)
        self._save_report(machine_id, report_type, start_dt, end_dt, machine, stats, alerts, generated_by)
        return buffer.getvalue()

    # ── Excel Report ──────────────────────────────────────────
    def generate_excel(
        self,
        machine_id: int,
        report_type: str,
        start_dt: datetime,
        end_dt: datetime,
        generated_by: int | None = None,
    ) -> bytes:
        """Generate an Excel report and return as bytes."""
        machine   = execute_one("SELECT * FROM machines WHERE id = %s", (machine_id,))
        if not machine:
            raise ValueError(f"Machine {machine_id} not found")

        stats     = self._get_stats(machine_id, start_dt, end_dt)
        alerts    = self._get_alerts(machine_id, start_dt, end_dt)
        telemetry = self._get_telemetry(machine_id, start_dt, end_dt)

        wb = openpyxl.Workbook()

        # ── Sheet 1: Summary ─────────────────────────────────
        ws = wb.active
        ws.title = "Summary"

        blue_fill  = PatternFill("solid", fgColor="0057A8")
        red_fill   = PatternFill("solid", fgColor="E31837")
        light_fill = PatternFill("solid", fgColor="F1F5F9")
        white_font = Font(color="FFFFFF", bold=True, size=11)
        bold_font  = Font(bold=True)
        blue_font  = Font(color="0057A8", bold=True, size=14)

        # Title
        ws.merge_cells('A1:F1')
        ws['A1'] = f'EDGEAI — {machine["machine_name"]} {report_type.upper()} REPORT'
        ws['A1'].font = Font(color="FFFFFF", bold=True, size=14)
        ws['A1'].fill = blue_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 35

        ws.merge_cells('A2:F2')
        ws['A2'] = f'{machine["machine_type"]} · {machine["location"]} · {start_dt.strftime("%d %b %Y")} to {end_dt.strftime("%d %b %Y")}'
        ws['A2'].font = Font(color="64748B", size=10)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[2].height = 20

        # Key metrics row
        ws.row_dimensions[4].height = 25
        for col, (label, value) in enumerate([
            ('Total Readings', int(stats.get('reading_count') or 0)),
            ('Anomalies', int(stats.get('anomaly_count') or 0)),
            ('Uptime %', self._calc_uptime(machine_id, start_dt, end_dt)),
            ('Avg Temp °C', round(_safe_float(stats.get('avg_temp')), 1)),
        ], start=1):
            ws.cell(row=4, column=col, value=label).font = Font(bold=True, size=9, color="64748B")
            ws.cell(row=4, column=col).fill = light_fill
            ws.cell(row=5, column=col, value=value).font = blue_font
            ws.cell(row=5, column=col).alignment = Alignment(horizontal='center')

        # Stats table
        ws['A7'] = 'PERFORMANCE STATISTICS'
        ws['A7'].font = Font(bold=True, size=11, color="0057A8")

        headers = ['Parameter', 'Average', 'Maximum', 'Minimum', 'Unit', 'Status']
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=8, column=col, value=h)
            cell.font = white_font
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal='center')

        rows = [
            ['Temperature', round(_safe_float(stats.get('avg_temp')),2), round(_safe_float(stats.get('max_temp')),2), round(_safe_float(stats.get('min_temp')),2), '°C', 'WARNING' if _safe_float(stats.get('max_temp')) > 95 else 'NORMAL'],
            ['Vibration',   round(_safe_float(stats.get('avg_vibration')),4), round(_safe_float(stats.get('max_vibration')),4), '—', 'mm/s', 'WARNING' if _safe_float(stats.get('max_vibration')) > 0.4 else 'NORMAL'],
            ['RPM',         round(_safe_float(stats.get('avg_rpm')),1), '—', '—', 'rpm', 'NORMAL'],
            ['Power',       round(_safe_float(stats.get('avg_power')),2), round(_safe_float(stats.get('max_power')),2), '—', 'W', 'WARNING' if _safe_float(stats.get('max_power')) > 1500 else 'NORMAL'],
        ]
        for r, row in enumerate(rows, start=9):
            for c, val in enumerate(row, start=1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.fill = light_fill if r % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
                if c == 6:
                    cell.font = Font(color="DC2626" if val == 'WARNING' else "059669", bold=True)

        # Column widths
        for col, width in zip('ABCDEF', [20, 15, 15, 15, 10, 12]):
            ws.column_dimensions[chr(64+col if isinstance(col, int) else ord(col))].width = width

        # ── Sheet 2: Telemetry Data ───────────────────────────
        ws2 = wb.create_sheet("Telemetry Data")
        tel_headers = ['Timestamp', 'Temperature (°C)', 'Vibration (mm/s)', 'RPM', 'Power (W)', 'Pressure (bar)', 'Anomaly']
        for col, h in enumerate(tel_headers, start=1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = white_font
            cell.fill = blue_fill
            cell.alignment = Alignment(horizontal='center')

        for r, row in enumerate(telemetry[:1000], start=2):
            ws2.cell(row=r, column=1, value=str(row.get('timestamp', ''))[:19])
            ws2.cell(row=r, column=2, value=round(_safe_float(row.get('temperature')), 2))
            ws2.cell(row=r, column=3, value=round(_safe_float(row.get('vibration')), 4))
            ws2.cell(row=r, column=4, value=round(_safe_float(row.get('rpm')), 1))
            ws2.cell(row=r, column=5, value=round(_safe_float(row.get('power_consumption')), 2))
            ws2.cell(row=r, column=6, value=round(_safe_float(row.get('pressure')), 2))
            ws2.cell(row=r, column=7, value='YES' if row.get('is_anomaly') else 'NO')
            if row.get('is_anomaly'):
                for col in range(1, 8):
                    ws2.cell(row=r, column=col).fill = PatternFill("solid", fgColor="FEF2F2")

        for col, width in enumerate([22, 18, 18, 12, 12, 14, 10], start=1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        # ── Sheet 3: Alerts ───────────────────────────────────
        ws3 = wb.create_sheet("Alerts")
        alert_headers = ['Timestamp', 'Alert Type', 'Severity', 'Message', 'Resolved']
        for col, h in enumerate(alert_headers, start=1):
            cell = ws3.cell(row=1, column=col, value=h)
            cell.font = white_font
            cell.fill = red_fill
            cell.alignment = Alignment(horizontal='center')

        for r, a in enumerate(alerts, start=2):
            ws3.cell(row=r, column=1, value=str(a.get('timestamp', ''))[:19])
            ws3.cell(row=r, column=2, value=str(a.get('alert_type', '')).replace('_', ' ').title())
            sev_cell = ws3.cell(row=r, column=3, value=str(a.get('severity', '')).upper())
            sev_cell.font = Font(color="DC2626" if a.get('severity') == 'critical' else "D97706", bold=True)
            ws3.cell(row=r, column=4, value=str(a.get('message', '')))
            ws3.cell(row=r, column=5, value='Yes' if a.get('is_resolved') else 'No')

        for col, width in enumerate([22, 25, 12, 60, 10], start=1):
            ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

        self._save_report(machine_id, report_type, start_dt, end_dt, machine, stats, alerts, generated_by)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ── Helpers ───────────────────────────────────────────────
    def _get_stats(self, machine_id, start_dt, end_dt):
        return execute_one(
            """
            SELECT COUNT(*) AS reading_count,
                ROUND(AVG(temperature)::numeric,2) AS avg_temp,
                ROUND(MAX(temperature)::numeric,2) AS max_temp,
                ROUND(MIN(temperature)::numeric,2) AS min_temp,
                ROUND(AVG(vibration)::numeric,4)   AS avg_vibration,
                ROUND(MAX(vibration)::numeric,4)   AS max_vibration,
                ROUND(AVG(rpm)::numeric,1)         AS avg_rpm,
                ROUND(AVG(power_consumption)::numeric,2) AS avg_power,
                ROUND(MAX(power_consumption)::numeric,2) AS max_power,
                SUM(CASE WHEN is_anomaly THEN 1 ELSE 0 END) AS anomaly_count
            FROM machine_data
            WHERE machine_id = %s AND timestamp BETWEEN %s AND %s
            """,
            (machine_id, start_dt, end_dt),
        ) or {}

    def _get_alerts(self, machine_id, start_dt, end_dt):
        return execute_many(
            """
            SELECT alert_type, message, severity, timestamp, is_resolved
            FROM alerts
            WHERE machine_id = %s AND timestamp BETWEEN %s AND %s
            ORDER BY timestamp DESC LIMIT 50
            """,
            (machine_id, start_dt, end_dt),
        )

    def _get_telemetry(self, machine_id, start_dt, end_dt):
        return execute_many(
            """
            SELECT timestamp, temperature, vibration, rpm, power_consumption, pressure, is_anomaly
            FROM machine_data
            WHERE machine_id = %s AND timestamp BETWEEN %s AND %s
            ORDER BY timestamp DESC LIMIT 1000
            """,
            (machine_id, start_dt, end_dt),
        )

    def _calc_uptime(self, machine_id, start_dt, end_dt) -> float:
        period_secs       = (end_dt - start_dt).total_seconds()
        expected_readings = period_secs / 2
        actual = execute_one(
            "SELECT COUNT(*) AS n FROM machine_data WHERE machine_id=%s AND timestamp BETWEEN %s AND %s",
            (machine_id, start_dt, end_dt),
        )
        return round(min(100, ((actual["n"] if actual else 0) / max(expected_readings, 1)) * 100), 2)

    def _rule_based_summary(self, stats, alerts, machine) -> str:
        anomalies = int(stats.get("anomaly_count") or 0)
        avg_temp  = _safe_float(stats.get("avg_temp"))
        max_vib   = _safe_float(stats.get("max_vibration"))
        lines = [f"Summary for {machine['machine_name']}:"]
        if anomalies == 0:
            lines.append("✅ No anomalies detected during this period.")
        else:
            lines.append(f"⚠ {anomalies} anomaly readings detected.")
        if avg_temp > 90:
            lines.append(f"🌡 Average temperature ({avg_temp}°C) is elevated — inspect cooling system.")
        if max_vib > 0.35:
            lines.append(f"📳 Peak vibration ({max_vib} mm/s) is high — check bearing condition.")
        critical = sum(1 for a in alerts if a.get("severity") == "critical")
        if critical:
            lines.append(f"🚨 {critical} critical alert(s) — immediate review required.")
        lines.append("Recommendation: Schedule routine inspection within 7 days.")
        return "\n".join(lines)

    def _save_report(self, machine_id, report_type, start_dt, end_dt, machine, stats, alerts, generated_by):
        content = {
            "machine_id":   machine_id,
            "machine_name": machine["machine_name"],
            "period":       {"start": start_dt.isoformat(), "end": end_dt.isoformat()},
            "stats":        {k: str(v) for k, v in (stats or {}).items()},
            "alert_count":  len(alerts),
        }
        title = f"{machine['machine_name']} — {report_type.capitalize()} Report ({start_dt.strftime('%d %b %Y')} – {end_dt.strftime('%d %b %Y')})"
        execute_write(
            "INSERT INTO reports (machine_id, report_type, title, content, generated_by) VALUES (%s, %s, %s, %s::jsonb, %s)",
            (machine_id, report_type, title, json.dumps(content), generated_by),
        )


def get_weekly_range():
    end = datetime.now(timezone.utc)
    return end - timedelta(days=7), end

def get_monthly_range():
    end = datetime.now(timezone.utc)
    return end - timedelta(days=30), end


report_generator = ReportGenerator()